"""
Synthetic industrial document generator for OpsBrain hackathon dataset.

Run once before the ingestion pipeline:
    python data/generate_synthetic.py

Outputs (all in data/raw/):
  PDFs   — compressor_c1_manual, valve_v12_maintenance_manual, pump_p3_oem_manual,
            sop_process_safety, sop_emergency_response
  Excel  — maintenance_log.xlsx  (18 work-order rows)
  Text   — incident_report_001..004.txt
  Image  — inspection_scan_001.png  (OCR demo: scanned-looking inspection report)

CONTRADICTIONS PLANTED (see data/CONTRADICTIONS.md for full detail):
  C1: Compressor-C1 max pressure — manual says 150 psi, SOP says 180 psi
  C2: Valve-V12 inspection interval — manual says 3 months, last log entry is 8.5 months ago
  C3: Pump-P3 emergency cooldown — manual says 15 min, emergency SOP says 5 min
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

RAW = Path(__file__).parent / "raw"
RAW.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _pdf(filepath: Path, title: str, content_blocks: list[tuple[str, str]]) -> None:
    """
    Render a PDF from a list of (style, text) tuples.
    Styles: 'title', 'h1', 'h2', 'body', 'bullet', 'warning', 'note', 'spacer'
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
    )

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=0.75 * inch,
    )

    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle("DocTitle", parent=base["Title"],
                                fontSize=18, alignment=TA_CENTER, spaceAfter=6),
        "subtitle": ParagraphStyle("Subtitle", parent=base["Normal"],
                                   fontSize=11, alignment=TA_CENTER,
                                   textColor=colors.grey, spaceAfter=12),
        "h1": ParagraphStyle("H1", parent=base["Heading1"],
                              fontSize=13, spaceBefore=14, spaceAfter=4,
                              textColor=colors.HexColor("#003366")),
        "h2": ParagraphStyle("H2", parent=base["Heading2"],
                              fontSize=11, spaceBefore=8, spaceAfter=4,
                              textColor=colors.HexColor("#004488")),
        "body": ParagraphStyle("Body", parent=base["BodyText"],
                               fontSize=10, spaceAfter=6, alignment=TA_JUSTIFY,
                               leading=14),
        "bullet": ParagraphStyle("Bullet", parent=base["BodyText"],
                                 fontSize=10, spaceAfter=3, leftIndent=18,
                                 firstLineIndent=-12),
        "warning": ParagraphStyle("Warn", parent=base["BodyText"],
                                  fontSize=10, spaceAfter=6,
                                  textColor=colors.HexColor("#8B0000"),
                                  leftIndent=12, rightIndent=12),
        "note": ParagraphStyle("Note", parent=base["BodyText"],
                               fontSize=10, spaceAfter=6,
                               textColor=colors.HexColor("#1a3c00"),
                               leftIndent=12, rightIndent=12),
        "small": ParagraphStyle("Small", parent=base["Normal"],
                                fontSize=8, textColor=colors.grey, spaceAfter=2),
    }

    story = []
    story.append(Paragraph(title, styles["title"]))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#003366")))
    story.append(Spacer(1, 0.15 * inch))

    for style_key, text in content_blocks:
        if style_key == "spacer":
            story.append(Spacer(1, float(text) * inch))
        elif style_key == "hr":
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        elif style_key == "pagebreak":
            story.append(PageBreak())
        elif style_key == "warning":
            story.append(Paragraph(f"[WARNING] {text}", styles["warning"]))
        elif style_key == "note":
            story.append(Paragraph(f"[NOTE] {text}", styles["note"]))
        elif style_key in styles:
            story.append(Paragraph(text, styles[style_key]))
        else:
            story.append(Paragraph(text, styles["body"]))

    doc.build(story)
    print(f"  Created: {filepath.name}")


# ---------------------------------------------------------------------------
# PDF 1 — Compressor-C1 OEM Manual
# PLANTS: max operating pressure = 150 psi  (SOP says 180 psi → Contradiction #2)
# ---------------------------------------------------------------------------

def gen_compressor_c1_manual() -> None:
    _pdf(
        RAW / "compressor_c1_manual.pdf",
        "CENTRIFUGAL COMPRESSOR — OPERATION & MAINTENANCE MANUAL",
        [
            ("subtitle", "Equipment Tag: Compressor-C1 | Model: CX-400 Series | Rev. 4 (2022)"),
            ("small", "IndustraTech Industries Ltd. | Document No: ITI-OM-CX400-001 | CONFIDENTIAL"),
            ("spacer", "0.1"),

            ("h1", "SECTION 1 — EQUIPMENT OVERVIEW"),
            ("body",
             "This manual governs operation, preventive maintenance, and safe shutdown of the "
             "Compressor-C1 centrifugal compressor unit installed in Process Area B, Bay 3 of "
             "the facility. It supersedes all previous revisions. All personnel performing "
             "maintenance must read and acknowledge this document before commencing work on "
             "Compressor-C1."),
            ("body",
             "Equipment Tag: Compressor-C1 | Type: Single-stage centrifugal compressor | "
             "Model: CX-400 Series | Serial No: CC-2019-00147 | "
             "Manufacturer: IndustraTech Industries Ltd. | Installation Date: 2019-03-15 | "
             "Design Life: 20 years"),

            ("h1", "SECTION 2 — TECHNICAL SPECIFICATIONS"),
            ("body",
             "The following parameters define the design limits and normal operating envelope "
             "for Compressor-C1. Do NOT operate outside these limits without written "
             "authorisation from the Plant Engineering Department."),
            ("body", "Rated Speed: 3,550 RPM"),
            ("body", "Suction Pressure (normal): 10 – 25 psi"),
            ("body", "Discharge Pressure (normal): 120 – 145 psi"),
            ("body", "Maximum Operating Pressure: 150 psi"),
            ("body", "Design Pressure: 165 psi"),
            ("body", "Normal Operating Temperature: 180 °F to 220 °F"),
            ("body", "Maximum Outlet Temperature: 280 °F"),
            ("body", "Motor Power Rating: 250 kW"),
            ("body", "Shaft Seal Type: Mechanical (Type III API 682)"),
            ("body", "Impeller Diameter: 14 inches | Material: 316 Stainless Steel"),
            ("body", "Bearing Type: Hydrodynamic oil-film bearings, ISO VG 46 lubricant"),
            ("warning",
             "The Maximum Operating Pressure for Compressor-C1 is 150 psi. This value is "
             "determined by the OEM based on the structural limits of the impeller, volute "
             "casing, and shaft seals. Sustained operation above 150 psi risks catastrophic "
             "seal failure and potential process gas release. This limit may NOT be revised "
             "without a formal Management of Change (MOC) review and OEM written concurrence."),

            ("pagebreak", ""),

            ("h1", "SECTION 3 — PREVENTIVE MAINTENANCE SCHEDULE"),
            ("h2", "3.1 Daily Operator Checks"),
            ("bullet", "• Verify suction and discharge pressures within normal operating range."),
            ("bullet", "• Check bearing oil level and bearing temperature indicators (normal: < 200 °F)."),
            ("bullet", "• Listen for abnormal noise or vibration (baseline: < 3.5 mm/s RMS)."),
            ("bullet", "• Record all readings in the Daily Operations Log (Form OPS-DL-001)."),

            ("h2", "3.2 Semi-Annual Inspection (6-Month Interval — Certified Technician Required)"),
            ("bullet", "• Disassemble and inspect impeller for erosion, fouling, and cavitation pitting."),
            ("bullet", "• Replace mechanical shaft seal if wear exceeds 0.5 mm on any sealing face."),
            ("bullet", "• Flush and replace bearing lubrication oil (ISO VG 46); inspect oil for metallic particles."),
            ("bullet", "• Inspect and clean suction filter elements; replace if differential pressure > 2 psi."),
            ("bullet", "• Verify shaft alignment to within 0.002 in TIR using laser alignment tool."),
            ("bullet", "• Complete PM checklist Form EQ-PM-009 and submit to Maintenance Supervisor."),

            ("h2", "3.3 Annual Overhaul"),
            ("bullet", "• Full disassembly and dimensional inspection of all rotating parts."),
            ("bullet", "• Non-destructive testing (NDT) of impeller and pressure casing."),
            ("bullet", "• Replace all seals, gaskets, wear rings, and O-rings regardless of condition."),
            ("bullet", "• Alignment certification required from a certified vibration analyst."),
            ("bullet", "• Pressure hydrostatic test to 1.5 × design pressure (247.5 psi) after reassembly."),

            ("h1", "SECTION 4 — OPERATING PROCEDURES"),
            ("h2", "4.1 Normal Start Sequence"),
            ("bullet", "1. Confirm lube oil system is pressurised to minimum 18 psi supply pressure."),
            ("bullet", "2. Open suction isolation valve V-101 slowly over 30 seconds to full open."),
            ("bullet", "3. Start drive motor per Electrical SOP SOP-ELEC-003."),
            ("bullet", "4. Monitor startup current — must return to nameplate amps within 15 seconds."),
            ("bullet", "5. Verify discharge pressure builds to 120–145 psi within 60 seconds of start."),
            ("bullet", "6. Log startup time and initial readings in the Daily Operations Log."),

            ("h2", "4.2 Normal Shutdown"),
            ("bullet", "1. Gradually reduce process load over 5 minutes minimum."),
            ("bullet", "2. Close discharge isolation valve V-102."),
            ("bullet", "3. De-energise motor at MCC Panel B-12."),
            ("bullet", "4. Allow unit to coast to rest — do NOT apply mechanical braking."),
            ("bullet", "5. Close suction isolation valve V-101 after unit reaches complete stop."),
            ("bullet", "6. Log shutdown time and reason in the operations log."),

            ("h2", "4.3 Emergency Shutdown (ESD)"),
            ("warning",
             "If discharge pressure exceeds 150 psi, bearing temperature exceeds 250 °F, "
             "excessive vibration is detected, or abnormal noise occurs: press ESD pushbutton "
             "ESD-C1 immediately. Notify control room and shift supervisor. Do NOT attempt "
             "restart without written Engineering clearance."),
        ],
    )


# ---------------------------------------------------------------------------
# PDF 2 — Valve-V12 Maintenance Manual
# PLANTS: inspection interval = every 3 months / 90 days
#         (maintenance log shows last inspection 2025-11-05 → Contradiction #1 decay)
# ---------------------------------------------------------------------------

def gen_valve_v12_manual() -> None:
    _pdf(
        RAW / "valve_v12_maintenance_manual.pdf",
        "GATE VALVE MAINTENANCE MANUAL — VALVE-V12",
        [
            ("subtitle", "Equipment Tag: Valve-V12 | Type: Motor-Operated Gate Valve | Rev. 2 (2021)"),
            ("small", "PipeFlow Engineering Corp | Document No: PFE-MNT-GV-012 | CONTROLLED COPY"),
            ("spacer", "0.1"),

            ("h1", "SECTION 1 — EQUIPMENT IDENTIFICATION"),
            ("body",
             "This document provides maintenance, inspection, and operational guidance for "
             "Valve-V12, a DN150 motor-operated gate valve installed on the high-pressure "
             "process gas supply line in Process Area B. Valve-V12 is a safety-critical "
             "isolation device. Deviation from this maintenance schedule must be approved "
             "by the Plant Safety Engineer in writing."),
            ("body",
             "Equipment Tag: Valve-V12 | Type: Motor-Operated Gate Valve | "
             "Size: DN150 (6 inch) | Pressure Rating: ANSI Class 300 (740 psi CWP) | "
             "Body Material: Carbon Steel ASTM A216 WCB | "
             "Actuation: Electric motor actuator, 480V 3-phase | "
             "Installed: 2018-06-20 | Serial No: PFE-GV-2018-0512"),

            ("h1", "SECTION 2 — TECHNICAL SPECIFICATIONS"),
            ("body", "Maximum Allowable Working Pressure (MAWP): 740 psi at 100 °F"),
            ("body", "Design Temperature Range: –20 °F to +450 °F"),
            ("body", "Seat Leakage Class: ANSI/FCI 70-2 Class IV (0.01% of rated capacity)"),
            ("body", "Stem Diameter: 50 mm | Stem Material: AISI 316 Stainless Steel"),
            ("body", "Actuator Torque Output: 900 N·m | Actuator Duty: S2-15 min"),
            ("body", "Open/Close Travel Time: 45 seconds at rated torque"),
            ("body", "Packing Type: Graphite chevron rings (5-ring set), PN-GR-150"),

            ("h1", "SECTION 3 — INSPECTION SCHEDULE"),
            ("warning",
             "Valve-V12 is a SAFETY-CRITICAL isolation valve on a high-pressure process line. "
             "The inspection interval MUST NOT be extended without a formal risk assessment "
             "signed by the Plant Safety Engineer and the Operations Manager."),
            ("h2", "3.1 Required Inspection Frequency"),
            ("body",
             "Valve-V12 shall be inspected every 3 months (90 days). This interval is "
             "mandated by the OEM based on the valve's service conditions (high-pressure "
             "cycling, elevated temperature, potential for packing degradation). "
             "Failure to comply with this schedule voids the OEM warranty and constitutes "
             "a documented safety deviation that must be reported to the HSE Department."),
            ("h2", "3.2 Quarterly Inspection Checklist (every 90 days)"),
            ("bullet", "• Perform open/close cycle test — verify full travel within 45 ± 5 seconds."),
            ("bullet", "• Inspect stem packing for weeping or visible leakage; replace if any leakage detected."),
            ("bullet", "• Torque packing gland nuts to 35 N·m ± 5 N·m (do not over-torque)."),
            ("bullet", "• Lubricate stem threads with approved thread compound (Molykote P-37 or equivalent)."),
            ("bullet", "• Verify actuator limit switches — confirm open and closed position indication."),
            ("bullet", "• Check actuator motor insulation resistance (minimum 1 MΩ at 500V DC)."),
            ("bullet", "• Inspect body for external corrosion, coating damage, and mechanical damage."),
            ("bullet", "• Record all findings in Work Order and Valve Inspection Report Form EQ-VI-012."),

            ("h2", "3.3 Annual Inspection (in addition to quarterly)"),
            ("bullet", "• Full disassembly and internal inspection of gate, seat rings, and body cavity."),
            ("bullet", "• Replace packing set and seat rings regardless of apparent condition."),
            ("bullet", "• Actuator gearbox oil change (Shell Omala S2 G 220 or equivalent)."),
            ("bullet", "• Hydrostatic shell test at 1,110 psi (1.5 × MAWP) for 10 minutes."),

            ("h1", "SECTION 4 — MAINTENANCE PROCEDURES"),
            ("h2", "4.1 Packing Replacement Procedure"),
            ("bullet", "1. Isolate valve from process — confirm upstream and downstream blind flanges are installed."),
            ("bullet", "2. De-energise actuator at local disconnect; lock out and tag out (LOTO Procedure SAFE-LOTO-007)."),
            ("bullet", "3. Remove packing gland follower and extract worn packing rings."),
            ("bullet", "4. Clean stem and stuffing box bore with lint-free cloth; inspect for scoring."),
            ("bullet", "5. Install new graphite rings (PN-GR-150) — stagger cuts by 90° minimum."),
            ("bullet", "6. Replace gland follower and torque gland nuts to 25 N·m (initial)."),
            ("bullet", "7. Re-pressurise and perform packing leak test; tighten to 35 N·m if weeping persists."),
            ("bullet", "8. Restore power and perform full open/close cycle test. Document in work order."),
        ],
    )


# ---------------------------------------------------------------------------
# PDF 3 — Pump-P3 OEM Manual
# PLANTS: emergency shutdown cooling period = 15 minutes
#         (Emergency SOP says 5 minutes → Contradiction #3)
# ---------------------------------------------------------------------------

def gen_pump_p3_manual() -> None:
    _pdf(
        RAW / "pump_p3_oem_manual.pdf",
        "CENTRIFUGAL PUMP — OEM OPERATION & MAINTENANCE MANUAL",
        [
            ("subtitle", "Equipment Tag: Pump-P3 | Model: FlowMaster FP-200 | Rev. 3 (2023)"),
            ("small", "FlowDynamics Systems Inc. | Document No: FDS-OM-FP200-P3 | CONTROLLED COPY"),
            ("spacer", "0.1"),

            ("h1", "SECTION 1 — EQUIPMENT OVERVIEW"),
            ("body",
             "This manual provides complete operating and maintenance instructions for Pump-P3, "
             "a horizontal end-suction centrifugal pump serving the cooling water circulation "
             "circuit in Plant Area C. Pump-P3 handles process cooling water at elevated "
             "temperatures and must be maintained in accordance with this document to ensure "
             "safe and reliable operation."),
            ("body",
             "Equipment Tag: Pump-P3 | Type: Horizontal end-suction centrifugal pump | "
             "Model: FlowMaster FP-200 | Serial No: FDS-2020-P3-0089 | "
             "Manufacturer: FlowDynamics Systems Inc. | Commissioned: 2020-09-01"),

            ("h1", "SECTION 2 — TECHNICAL SPECIFICATIONS"),
            ("body", "Rated Flow: 450 GPM at BEP"),
            ("body", "Rated Head: 120 ft (52 psi)"),
            ("body", "Maximum Operating Pressure (casing): 200 psi"),
            ("body", "Design Temperature: 250 °F maximum"),
            ("body", "Motor: 30 kW, 1,780 RPM, TEFC 480V 3-phase"),
            ("body", "Impeller Type: Closed, 316 SS | Diameter: 11.5 inches"),
            ("body", "Seal Type: Single mechanical seal (Flowserve Type 1 equivalent)"),
            ("body", "Bearing Lubrication: Grease-packed anti-friction bearings (Mobil Polyrex EM)"),
            ("body", "NPSHr at rated flow: 12 ft"),
            ("body", "Coupling: Flexible disc type, spacer design for back-pull-out"),

            ("h1", "SECTION 3 — PREVENTIVE MAINTENANCE"),
            ("h2", "3.1 Monthly Checks"),
            ("bullet", "• Check bearing temperature (normal: < 180 °F; alarm: 210 °F; trip: 230 °F)."),
            ("bullet", "• Inspect mechanical seal for leakage (allowable: < 5 drops/minute)."),
            ("bullet", "• Check vibration (normal: < 2.5 mm/s; investigate: > 4.5 mm/s RMS)."),
            ("bullet", "• Grease bearings: 15 grams Mobil Polyrex EM per bearing (once per 3 months)."),

            ("h2", "3.2 Annual Overhaul"),
            ("bullet", "• Replace mechanical seal assembly regardless of condition."),
            ("bullet", "• Inspect and measure impeller wear ring clearances; replace if > 0.030 in."),
            ("bullet", "• Replace bearing housing seals and repack bearings."),
            ("bullet", "• Perform hydraulic performance test and compare to original curve."),

            ("pagebreak", ""),

            ("h1", "SECTION 4 — EMERGENCY PROCEDURES"),
            ("h2", "4.1 Emergency Shutdown Triggers"),
            ("bullet", "• Bearing temperature exceeds 230 °F (auto-trip via TT-P3-001)."),
            ("bullet", "• Vibration exceeds 7.5 mm/s RMS (auto-trip via VT-P3-001)."),
            ("bullet", "• Mechanical seal failure (visible fluid loss > 50 ml/min)."),
            ("bullet", "• Suction pressure drops below 5 psi (cavitation risk — manual ESD)."),
            ("bullet", "• Any abnormal noise, smoke, or burning smell."),

            ("h2", "4.2 Emergency Shutdown Procedure and Post-Trip Access"),
            ("warning",
             "CRITICAL SAFETY REQUIREMENT — Following emergency shutdown of Pump-P3, a mandatory "
             "cooling period of 15 minutes minimum must elapse before any maintenance personnel "
             "access the pump casing, mechanical seal area, or bearing housing. "
             "This is required because Pump-P3 handles process water at up to 250 °F; residual "
             "heat in the casing and trapped pressurised fluid create burn and pressure-release "
             "hazards for up to 15 minutes after shutdown. "
             "This 15-minute cooling period is NON-NEGOTIABLE and may not be reduced by any "
             "field supervisor without written authorisation from the Plant Safety Manager. "
             "Failure to comply is a SERIOUS SAFETY VIOLATION."),
            ("bullet", "1. Press red ESD pushbutton P3-ESD-001 (or de-energise at MCC Panel C-08)."),
            ("bullet", "2. Close suction isolation valve P3-V-001 immediately after trip."),
            ("bullet", "3. Close discharge check valve P3-V-002 if not auto-actuated."),
            ("bullet", "4. Post a physical barrier and DANGER tag on the pump for the 15-minute cooling period."),
            ("bullet", "5. After 15 minutes: verify casing temperature < 120 °F using infrared thermometer."),
            ("bullet", "6. Depressurise casing via vent/drain before removing any bolted connections."),
            ("bullet", "7. Document ESD event, time, cause, and post-trip inspection in Work Order System."),
        ],
    )


# ---------------------------------------------------------------------------
# PDF 4 — SOP: Process Safety
# PLANTS: Compressor-C1 relief valve = 180 psi
#         (manual says max pressure is 150 psi → Contradiction #2)
# ---------------------------------------------------------------------------

def gen_sop_process_safety() -> None:
    _pdf(
        RAW / "sop_process_safety.pdf",
        "STANDARD OPERATING PROCEDURE — PROCESS SAFETY MANAGEMENT",
        [
            ("subtitle", "Document No: SOP-PSM-004 | Rev. 5 | Effective: 2023-01-15"),
            ("small",
             "Plant Safety Department | Approved: Plant Manager | "
             "Review Due: 2025-01-15 | Based on OSHA 29 CFR 1910.119"),
            ("spacer", "0.1"),

            ("h1", "1. PURPOSE AND SCOPE"),
            ("body",
             "This Standard Operating Procedure (SOP) establishes the process safety management "
             "requirements for all rotating and pressure-containing equipment in Plant Areas A, B, "
             "and C. It implements the requirements of OSHA 29 CFR 1910.119 (Process Safety "
             "Management of Highly Hazardous Chemicals) and the facility's internal Process Safety "
             "Management (PSM) programme."),
            ("body",
             "This SOP covers: pressure relief device settings, compressor operating limits, "
             "pump and valve pressure testing, and management of change procedures for process "
             "equipment. All operations and maintenance personnel must be trained on this SOP "
             "annually. Training records must be retained for a minimum of 3 years."),

            ("h1", "2. GENERAL PRESSURE MANAGEMENT REQUIREMENTS"),
            ("body",
             "All pressure-containing equipment must be operated within the pressure limits "
             "specified in the applicable Mechanical Integrity programme and equipment datasheets. "
             "Pressure relief devices (PRDs) must be sized and set in accordance with API 520 "
             "and API 521. PRD settings must be reviewed and verified during each planned "
             "shutdown or at a maximum interval of 36 months."),
            ("body",
             "The following general principles apply to all pressure systems:"),
            ("bullet", "• No PRD shall be set above the Maximum Allowable Working Pressure (MAWP) of the protected vessel or piping."),
            ("bullet", "• PRD set points shall be documented in the Equipment Pressure Relief Register (Form EQ-PRR-001)."),
            ("bullet", "• Any deviation from registered PRD settings requires an approved MOC (Management of Change) before implementation."),
            ("bullet", "• PRD testing and certification shall be performed by a certified pressure testing contractor annually."),

            ("pagebreak", ""),

            ("h1", "3. EQUIPMENT-SPECIFIC PRESSURE SETTINGS"),
            ("h2", "3.1 General Compressor Relief Requirements"),
            ("body",
             "All compressors in Process Areas A, B, and C must be protected by a dedicated "
             "pressure safety valve (PSV) mounted on the high-pressure discharge line. PSV "
             "settings must comply with Process Engineering specification PE-2021-047 and must "
             "be verified during each annual turnaround."),

            ("h2", "3.2 Compressor-C1 Relief Valve Settings"),
            ("body",
             "Compressor-C1 is protected by PSV-C1-001 (3-inch × 4-inch full-nozzle PSV, "
             "Crosby Series HCI, spring code R-14). In accordance with Process Engineering "
             "specification PE-2021-047 and the facility's process hazard analysis (PHA) "
             "conducted in 2021, the PSV set point for Compressor-C1 is established as follows:"),
            ("body",
             "Compressor-C1 pressure relief valve shall be set to 180 psi maximum allowable "
             "working pressure (MAWP) as per Process Engineering specification PE-2021-047. "
             "This set point accounts for the downstream piping MAWP rating of 200 psi "
             "(ANSI Class 150 carbon steel) and provides a 10% margin above the normal "
             "maximum discharge pressure of 163 psi observed during high-load operations."),
            ("warning",
             "The PSV set point of 180 psi for Compressor-C1 was established based on a "
             "2021 process hazard analysis. Any proposal to change this set point requires "
             "a full MOC review with PHA sign-off, not just equipment-level approval. "
             "Contact the Process Safety Engineer before adjusting PSV-C1-001."),

            ("h2", "3.3 Pump Relief Settings"),
            ("body",
             "Pump-P3 discharge line is protected by PSV-P3-001 set at 220 psi. "
             "All other pump PSVs in Plant Area C are set at 210 psi unless otherwise "
             "specified in the Equipment Pressure Relief Register."),

            ("h1", "4. MANAGEMENT OF CHANGE PROCEDURE"),
            ("body",
             "Any change to process safety parameters — including PRD set points, operating "
             "pressure limits, equipment isolation schemes, or safety interlock settings — "
             "must follow the facility's Management of Change (MOC) procedure (SOP-MOC-001) "
             "before implementation. MOC review must include: process hazard analysis update, "
             "P&ID mark-up, operations and maintenance notification, and training completion "
             "verification."),
            ("body",
             "Unauthorised changes to PSV settings or operating pressure limits are a "
             "SERIOUS SAFETY VIOLATION and will result in immediate investigation under "
             "the facility's discipline and incident reporting procedures."),

            ("h1", "5. RECORD KEEPING"),
            ("bullet", "• Equipment Pressure Relief Register (EQ-PRR-001): maintained by Mechanical Integrity team."),
            ("bullet", "• PSV test certificates: retained for a minimum of the life of the equipment plus 5 years."),
            ("bullet", "• MOC records: retained for a minimum of 10 years."),
            ("bullet", "• Operator training records: retained for a minimum of 3 years."),
        ],
    )


# ---------------------------------------------------------------------------
# PDF 5 — SOP: Emergency Response
# PLANTS: Pump-P3 cooldown = 5 minutes
#         (OEM manual says 15 minutes → Contradiction #3)
# ---------------------------------------------------------------------------

def gen_sop_emergency_response() -> None:
    _pdf(
        RAW / "sop_emergency_response.pdf",
        "STANDARD OPERATING PROCEDURE — EMERGENCY RESPONSE",
        [
            ("subtitle", "Document No: SOP-ERP-002 | Rev. 3 | Effective: 2024-03-01"),
            ("small",
             "Operations Department | Approved: Operations Manager | "
             "Review Due: 2026-03-01 | Aligned with OSHA 29 CFR 1910.38"),
            ("spacer", "0.1"),

            ("h1", "1. PURPOSE AND SCOPE"),
            ("body",
             "This SOP defines emergency response actions for unplanned equipment shutdowns, "
             "process upsets, and safety-critical events in Plant Areas A, B, and C. It provides "
             "first-response guidance to operators and technicians before specialist teams arrive. "
             "This document must be accessible at all operator workstations and field control "
             "panels. All operations staff must be trained and competency-assessed on this "
             "procedure annually."),

            ("h1", "2. GENERAL EMERGENCY RESPONSE PRINCIPLES"),
            ("body",
             "On detection of any emergency condition: STOP — THINK — ACT. Do not rush. "
             "A hasty response to an equipment emergency creates additional hazards. "
             "Assess the situation, ensure your own safety first, then initiate the "
             "appropriate response from this SOP."),
            ("bullet", "• Personal safety is the first priority. Do not approach energised or hot equipment without PPE."),
            ("bullet", "• Activate the local ESD system and notify the control room before any other action."),
            ("bullet", "• Establish an exclusion zone of minimum 5 metres around any tripped rotating equipment."),
            ("bullet", "• Do not attempt to restart tripped equipment without shift supervisor authorisation."),
            ("bullet", "• Document all emergency events in the Shift Log within 30 minutes of occurrence."),

            ("pagebreak", ""),

            ("h1", "3. EQUIPMENT-SPECIFIC EMERGENCY PROCEDURES"),
            ("h2", "3.1 Pump-P3 — Emergency Trip Response"),
            ("body",
             "Pump-P3 is equipped with automatic trip protection on high bearing temperature "
             "(230 °F), high vibration (7.5 mm/s), and low suction pressure (< 5 psi). "
             "When Pump-P3 trips, the following response procedure applies:"),
            ("bullet", "1. Confirm trip annunciation at DCS and local panel (Panel C-08 red trip lamp)."),
            ("bullet", "2. Notify control room operator and shift supervisor immediately."),
            ("bullet", "3. Identify and record the trip cause from DCS alarm history (Form OPS-EV-002)."),
            ("bullet", "4. Establish exclusion zone; place CAUTION barrier around Pump-P3."),
            ("body",
             "For Pump-P3 post-trip access: a 5-minute cooldown period is sufficient for "
             "routine post-trip inspection checks by a qualified technician wearing appropriate "
             "PPE (heat-resistant gloves, face shield, and safety footwear). After 5 minutes, "
             "the technician may approach to check the bearing temperature gauge, inspect "
             "for visible seal leakage, and visually assess the pump for mechanical damage."),
            ("note",
             "The 5-minute cooldown applies to EXTERNAL VISUAL INSPECTION ONLY. Any "
             "maintenance work requiring removal of bolted connections, drain plugs, or "
             "access covers requires a full lock-out/tag-out (LOTO) per SAFE-LOTO-007 "
             "and must be planned as a scheduled maintenance activity, not an emergency response."),
            ("bullet", "5. Complete post-trip inspection report (Form EQ-PT-P3-001)."),
            ("bullet", "6. Consult Pump-P3 OEM manual for detailed root-cause guidance if bearing trip."),
            ("bullet", "7. Restart only with Shift Supervisor written authorisation after clearing root cause."),

            ("h2", "3.2 Compressor-C1 — Emergency Trip Response"),
            ("body",
             "Compressor-C1 trips automatically on high discharge pressure (> 150 psi), "
             "high bearing temperature (> 250 °F), or high vibration (> 6.0 mm/s RMS). "
             "Following a trip, maintain a 10-metre exclusion zone. Do not approach until "
             "the unit has come to a complete stop (typically 3–5 minutes coast-down). "
             "Notify Engineering before any restart attempt."),

            ("h2", "3.3 Valve-V12 — Actuator Failure Response"),
            ("body",
             "If Valve-V12 fails to respond to open/close commands: switch to manual override "
             "on the actuator handwheel (clockwise = close). Do not force the handwheel if "
             "resistance exceeds 150 N·m — escalate to Mechanical Maintenance. "
             "Log the failure in the Valve Inspection Register."),

            ("h1", "4. ESCALATION AND NOTIFICATION MATRIX"),
            ("body",
             "All emergency events must be escalated per the following matrix. Time targets "
             "are from the moment the event is confirmed, not from when it is first detected."),
            ("bullet", "• Within 5 min: Control room and Shift Supervisor notified."),
            ("bullet", "• Within 15 min: Maintenance Supervisor notified if equipment repair required."),
            ("bullet", "• Within 30 min: Plant Manager notified for any ESD activation."),
            ("bullet", "• Within 1 hour: HSE Manager notified for any injury, fire, or release event."),
            ("bullet", "• Within 24 hours: Incident Report Form INC-001 submitted to HSE Department."),
        ],
    )


# ---------------------------------------------------------------------------
# Excel — maintenance_log.xlsx
# PLANTS: Valve-V12 last inspection = 2025-11-05 (Contradiction #1 — decay)
# ---------------------------------------------------------------------------

def gen_maintenance_log() -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Log"

    headers = [
        "work_order_id", "equipment_tag", "date", "technician",
        "finding", "action_taken",
    ]

    rows = [
        ("WO-2025-0001", "Compressor-C1", "2025-01-15", "Mike Johnson",
         "Routine semi-annual inspection; all parameters within spec",
         "Replaced suction filter element; impeller clearance measured at 0.015 in (within tolerance)"),
        ("WO-2025-0002", "Pump-P3", "2025-01-28", "Sarah Lee",
         "High bearing temperature alarm: TT-P3-001 reading 218 deg F (alarm at 210 deg F)",
         "Repacked bearings with Mobil Polyrex EM; bearing temp returned to 165 deg F after restart"),
        ("WO-2025-0003", "Valve-V12", "2025-02-10", "Tom Williams",
         "Quarterly 90-day inspection; slight packing weep observed on stem",
         "Re-torqued packing gland to 35 Nm; verified full open/close cycle in 43 sec; weep stopped"),
        ("WO-2025-0004", "Turbine-T1", "2025-02-25", "Mike Johnson",
         "Routine vibration check: reading 3.2 mm/s RMS (threshold 4.5 mm/s)",
         "Vibration within acceptable range; increased monitoring frequency to bi-weekly"),
        ("WO-2025-0005", "Pump-P3", "2025-03-20", "Sarah Lee",
         "Post-bearing replacement follow-up inspection",
         "Bearing temperature stable at 158 deg F; no abnormalities; cleared for normal operation"),
        ("WO-2025-0006", "Compressor-C1", "2025-04-10", "Tom Williams",
         "Discharge pressure trending high: recorded 147 psi during peak load (normal max: 145 psi)",
         "Verified PSV-C1-001 set point; adjusted suction throttle to reduce load; pressure returned to 138 psi"),
        ("WO-2025-0007", "Valve-V12", "2025-05-12", "Mike Johnson",
         "Quarterly 90-day inspection (3 months from WO-2025-0003)",
         "Lubricated stem with Molykote P-37; verified actuator limit switches; no anomalies found"),
        ("WO-2025-0008", "Turbine-T1", "2025-05-28", "Sarah Lee",
         "Annual overhaul due",
         "Full blade inspection completed; replaced drive-end bearing (part no TRB-440); alignment certified at 0.001 in TIR"),
        ("WO-2025-0009", "Pump-P3", "2025-06-15", "Tom Williams",
         "Elevated vibration reading 3.8 mm/s during monthly check",
         "Removed and cleaned impeller; fouling buildup removed; vibration returned to 2.1 mm/s"),
        ("WO-2025-0010", "Compressor-C1", "2025-07-01", "Mike Johnson",
         "Semi-annual 6-month inspection",
         "Impeller inspected — no erosion; shaft seal wear 0.2 mm (within 0.5 mm limit); oil flushed and replaced"),
        ("WO-2025-0011", "Turbine-T1", "2025-07-20", "Sarah Lee",
         "Routine check; oil sample taken for analysis",
         "Oil sample submitted to lab (Ref: OIL-2025-0044); results awaited"),
        ("WO-2025-0012", "Valve-V12", "2025-08-14", "Tom Williams",
         "Quarterly 90-day inspection (3 months from WO-2025-0007)",
         "Replaced 5-ring packing set (PN-GR-150); actuator gearbox oil checked — acceptable level; full cycle test passed"),
        ("WO-2025-0013", "Compressor-C1", "2025-09-05", "Mike Johnson",
         "UNPLANNED: High vibration alarm triggered ESD-C1; vibration peaked at 8.2 mm/s RMS",
         "Root cause: impeller nut backing off. Impeller removed, nut re-torqued to 280 Nm with Loctite 243; restart approved by Engineering"),
        ("WO-2025-0014", "Pump-P3", "2025-09-22", "Sarah Lee",
         "Monthly routine check",
         "All parameters normal: bearing 162 deg F, vibration 2.3 mm/s, seal dry — no action required"),
        ("WO-2025-0015", "Turbine-T1", "2025-10-10", "Tom Williams",
         "UNPLANNED: Vibration exceeded threshold at 5.8 mm/s RMS — emergency work order raised",
         "Emergency realignment performed; balance weights adjusted; vibration reduced to 2.8 mm/s; cleared for operation"),
        ("WO-2025-0016", "Compressor-C1", "2025-10-30", "Mike Johnson",
         "Post-ESD inspection following WO-2025-0013 restart",
         "All parameters stable after restart; vibration 2.9 mm/s; discharge pressure 132 psi; cleared for normal operation"),
        ("WO-2025-0031", "Valve-V12", "2025-11-05", "Mike Johnson",
         "Quarterly 90-day inspection (3 months from WO-2025-0012)",
         "Replaced packing gland (preventive); lubricated stem assembly; full cycle test 44 sec — within spec"),
        ("WO-2026-0001", "Compressor-C1", "2026-01-10", "Tom Williams",
         "Annual overhaul; full disassembly",
         "Impeller replaced (wear beyond tolerance); all seals and gaskets renewed; hydrostatic test passed at 247 psi for 30 min"),
        ("WO-2026-0002", "Turbine-T1", "2026-02-15", "Mike Johnson",
         "Routine check; oil change due",
         "Oil changed (ISO VG 46); bearing temp 170 deg F; vibration 2.5 mm/s — normal condition"),
        ("WO-2026-0003", "Pump-P3", "2026-06-20", "Sarah Lee",
         "Monthly routine check",
         "Normal condition; bearing 155 deg F; vibration 2.2 mm/s; seal dry — no action required"),
    ]

    # Styling
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    normal_font = Font(name="Calibri", size=10)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(rows, start=2):
        ws.append(list(row))
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for cell in ws[i]:
            cell.font = normal_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = [16, 16, 12, 16, 55, 75]
    for col, width in zip("ABCDEF", col_widths):
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    path = RAW / "maintenance_log.xlsx"
    wb.save(str(path))
    print(f"  Created: {path.name}")


# ---------------------------------------------------------------------------
# Incident reports (text files)
# ---------------------------------------------------------------------------

def gen_incident_reports() -> None:
    reports = [
        (
            "incident_report_001.txt",
            """INCIDENT REPORT
Report No: IR-2025-0042
Date of Incident: 2025-09-05
Equipment Tag: Compressor-C1
Incident Type: Pressure / Vibration — ESD Activation
Severity: HIGH
Reported By: Mike Johnson
Location: Process Area B, Bay 3

DESCRIPTION OF INCIDENT:
At approximately 14:23 on 2025-09-05, the Emergency Shutdown System (ESD-C1) for Compressor-C1
activated automatically due to high vibration (8.2 mm/s RMS, trip set-point 6.0 mm/s RMS). The
unit tripped cleanly and coasted to rest within 4 minutes. No process gas release occurred.

The discharge pressure at the time of the trip was 143 psi — within the normal operating range
of 120–145 psi, and below the Maximum Operating Pressure of 150 psi specified in the OEM manual.

ROOT CAUSE:
Investigation found that the impeller nut had backed off due to inadequate torque retention. The
loose nut caused progressive impeller imbalance over several operating cycles, resulting in
increasing vibration until the trip threshold was reached.

CORRECTIVE ACTIONS TAKEN:
1. Impeller removed, inspected (no damage found), and reinstalled.
2. Impeller nut torqued to 280 Nm with Loctite 243 threadlocker applied.
3. Restart approved by Plant Engineering after alignment verification.
4. Engineering review of impeller nut torque specifications — findings added to PM checklist EQ-PM-009.

LESSONS LEARNED:
Impeller nut torque retention must be verified at every semi-annual inspection using a calibrated
torque wrench. The use of a thread-locking compound (Loctite 243 or approved equivalent) is now
mandatory during reassembly. Field tightening by feel is insufficient for high-speed rotating
components on Compressor-C1.

STATUS: CLOSED — 2025-09-12
""",
        ),
        (
            "incident_report_002.txt",
            """INCIDENT REPORT
Report No: IR-2025-0018
Date of Incident: 2025-01-28
Equipment Tag: Pump-P3
Incident Type: High Bearing Temperature — Alarm
Severity: MEDIUM
Reported By: Sarah Lee
Location: Plant Area C, Cooling Water Circuit

DESCRIPTION OF INCIDENT:
At 09:47 on 2025-01-28, the bearing temperature alarm for Pump-P3 (TT-P3-001) activated at the
DCS. The alarm set-point is 210 °F; the temperature at the time of alarm was 218 °F. The pump
was not tripped automatically (auto-trip set at 230 °F). The shift operator immediately notified
the Maintenance department and isolated the non-critical load to reduce the pump duty.

An emergency work order (WO-2025-0002) was raised. Maintenance technician Sarah Lee attended
within 20 minutes of the alarm.

ROOT CAUSE:
Inspection of the drive-end bearing found that the grease had degraded and partially escaped
the bearing housing through a worn lip seal. The bearing was operating in a near-dry condition,
leading to elevated temperature due to increased friction.

CORRECTIVE ACTIONS TAKEN:
1. Pump-P3 shut down per normal shutdown procedure (no ESD required).
2. Drive-end and non-drive-end bearings repacked with fresh Mobil Polyrex EM grease.
3. Worn lip seals on both bearing housings replaced.
4. Bearing housing end caps inspected and resealed.
5. Pump restarted; bearing temperature returned to 158 °F within 30 minutes.

LESSONS LEARNED:
The bearing regreasing interval for Pump-P3 shall be increased from quarterly to monthly during
summer months (June–September) when ambient temperatures in Plant Area C exceed 35 °C. The
bearing lip seal condition must be specifically checked at each quarterly maintenance visit.

STATUS: CLOSED — 2025-02-05
""",
        ),
        (
            "incident_report_003.txt",
            """INCIDENT REPORT
Report No: IR-2025-0071
Date of Incident: 2025-11-20
Equipment Tag: Valve-V12
Incident Type: Minor Packing Leak — Near Miss
Severity: LOW
Reported By: Tom Williams
Location: Process Area B, High-Pressure Gas Supply Line

DESCRIPTION OF INCIDENT:
On 2025-11-20, during a routine area walk-down, the operator noticed a faint smell of process
gas near Valve-V12. Visual inspection revealed a slight weep from the stem packing gland —
approximately 2–3 drops per minute of condensate carrying dissolved gas.

The leak was classified as a minor near-miss (no ignition source nearby; gas concentration
measured at < 5% LEL using a handheld CGI). The valve was not isolated as the leak was within
acceptable operational limits.

NOTE ON INSPECTION STATUS:
At the time of this incident, the last formal quarterly inspection of Valve-V12 was recorded as
WO-2025-0031 on 2025-11-05 (15 days prior). The packing replacement performed in WO-2025-0031
was confirmed complete; however, the gland may not have been torqued sufficiently to allow for
initial packing consolidation.

ROOT CAUSE:
Packing consolidation weep — normal behaviour within the first 30 days of new packing installation.
However, given that Valve-V12 is on a high-pressure safety-critical line, any gas weep must be
reported and assessed.

CORRECTIVE ACTIONS TAKEN:
1. Packing gland re-torqued by Tom Williams (in-service adjustment, within approved procedure).
2. Leak test with soapy water solution — no bubbles observed after re-torquing.
3. CGI reading confirmed 0% LEL after 30 minutes.
4. Incident logged per HSE near-miss reporting procedure.

LESSONS LEARNED:
New packing installations on high-pressure gas valves should include a scheduled 7-day follow-up
check to verify packing consolidation and re-torque if necessary. This follow-up check is not
currently included in the standard PM checklist — recommend updating Form EQ-VI-012.

STATUS: CLOSED — 2025-11-22
""",
        ),
        (
            "incident_report_004.txt",
            """INCIDENT REPORT
Report No: IR-2025-0063
Date of Incident: 2025-10-10
Equipment Tag: Turbine-T1
Incident Type: High Vibration — Emergency Work Order
Severity: MEDIUM
Reported By: Tom Williams
Location: Power Generation Building, Turbine Hall

DESCRIPTION OF INCIDENT:
At 11:15 on 2025-10-10, vibration monitoring for Turbine-T1 (VT-T1-001) showed a reading of
5.8 mm/s RMS — above the investigation threshold of 4.5 mm/s and approaching the trip threshold
of 7.5 mm/s. The control room operator raised an emergency work order (WO-2025-0015) and notified
the Mechanical Maintenance supervisor.

The turbine was not tripped immediately as the vibration was trending slowly upward. The shift
supervisor made the decision to continue operation at reduced load while maintenance attended.

ROOT CAUSE:
Laser alignment check performed by Tom Williams found the turbine shaft misaligned by 0.008 in
TIR on the coupling (acceptable limit: 0.003 in TIR). The misalignment had likely developed
gradually due to thermal growth of the foundation grout, which showed signs of cracking at two
of the four hold-down bolt locations.

CORRECTIVE ACTIONS TAKEN:
1. Turbine load reduced to 60% while maintenance was performed (controlled operation under
   close monitoring — vibration remained below trip threshold during this period).
2. Shaft realigned using laser alignment tool; final reading 0.001 in TIR.
3. Balance weights adjusted on the coupling hub.
4. Vibration reduced to 2.8 mm/s RMS immediately following realignment.
5. Foundation grout cracking reported to Civil Engineering for scheduled repair (CR-2025-0047).

LESSONS LEARNED:
Foundation grout condition for Turbine-T1 must be visually inspected during each annual overhaul.
The 6-monthly baseline vibration trending must be reviewed to catch misalignment development
earlier. Turbine-T1 vibration data to be included in the monthly plant reliability meeting agenda.

STATUS: CLOSED — 2025-10-18
""",
        ),
    ]

    for filename, content in reports:
        path = RAW / filename
        path.write_text(content, encoding="utf-8")
        print(f"  Created: {filename}")


# ---------------------------------------------------------------------------
# Scanned inspection image (OCR demo)
# ---------------------------------------------------------------------------

def gen_scanned_inspection() -> None:
    import math
    import random

    import numpy as np
    from PIL import Image, ImageDraw, ImageFilter, ImageFont

    width, height = 850, 1100
    img = Image.new("L", (width, height), color=252)
    draw = ImageDraw.Draw(img)

    # Try a Windows system font; fall back to PIL default
    def _font(size: int):
        for path in [
            "C:/Windows/Fonts/cour.ttf",     # Courier New
            "C:/Windows/Fonts/arial.ttf",    # Arial
            "C:/Windows/Fonts/calibri.ttf",  # Calibri
        ]:
            try:
                return ImageFont.truetype(path, size)
            except OSError:
                continue
        return ImageFont.load_default()

    f_title = _font(20)
    f_head  = _font(14)
    f_body  = _font(12)
    f_small = _font(10)

    # --- Header ---
    draw.text((width // 2, 55), "EQUIPMENT INSPECTION REPORT", font=f_title, fill=10, anchor="mm")
    draw.line([(50, 75), (width - 50, 75)], fill=30, width=2)
    draw.line([(50, 79), (width - 50, 79)], fill=80, width=1)

    # --- Metadata block ---
    meta = [
        ("Report No:",        "IR-2026-0078"),
        ("Date of Inspection:", "2026-01-15"),
        ("Inspector Name:",   "John T. Smith, Cert. MI-2341"),
        ("Equipment Tag:",    "Compressor-C1"),
        ("Equipment Type:",   "Centrifugal Compressor, Single-Stage"),
        ("Location:",         "Process Area B, Bay 3"),
        ("Plant / Unit:",     "Gas Compression Facility — Train 2"),
    ]
    y = 100
    for label, value in meta:
        draw.text((60, y), label, font=f_head, fill=20)
        draw.text((260, y), value, font=f_head, fill=0)
        y += 28

    draw.line([(50, y + 5), (width - 50, y + 5)], fill=80, width=1)
    y += 20

    # --- Inspection findings ---
    draw.text((60, y), "INSPECTION FINDINGS", font=f_head, fill=10)
    y += 28

    findings = [
        ("1.", "VISUAL INSPECTION:", "No visible cracks, corrosion, or coating damage on pressure casing."),
        ("2.", "VIBRATION (baseline):", "2.9 mm/s RMS — within normal operating range (< 3.5 mm/s)."),
        ("3.", "BEARING TEMPERATURE:", "185 deg F — within OEM normal range (< 200 deg F)."),
        ("4.", "DISCHARGE PRESSURE:", "138 psi — within operating envelope (120-145 psi normal; max 150 psi per OEM)."),
        ("5.", "SUCTION FILTER dP:", "1.2 psi — within limit (replace at > 2.0 psi)."),
        ("6.", "SHAFT SEAL CONDITION:", "Mechanical seal dry; wear measured at 0.18 mm (limit: 0.5 mm)."),
        ("7.", "LUBE OIL SYSTEM:", "Oil level normal; no metallic particles on magnetic plug."),
        ("8.", "COUPLING ALIGNMENT:", "0.001 in TIR — within tolerance (limit: 0.002 in TIR)."),
    ]
    for num, category, detail in findings:
        draw.text((70, y), num, font=f_body, fill=0)
        draw.text((90, y), category, font=f_body, fill=0)
        draw.text((90, y + 16), detail, font=f_small, fill=40)
        y += 44

    draw.line([(50, y), (width - 50, y)], fill=80, width=1)
    y += 15

    # --- Overall result ---
    draw.text((60, y), "OVERALL RESULT:", font=f_head, fill=0)
    draw.text((220, y), "PASS — No defects found. Unit cleared for continued operation.", font=f_head, fill=0)
    y += 35

    draw.text((60, y), "Next scheduled inspection due: 2026-07-15 (6-month interval)", font=f_body, fill=30)
    y += 25
    draw.text((60, y), "PM Checklist Form: EQ-PM-009 completed and filed.", font=f_body, fill=30)
    y += 45

    # --- Signatures ---
    draw.line([(50, y), (width - 50, y)], fill=80, width=1)
    y += 15
    draw.text((60,  y), "Inspector Signature:", font=f_body, fill=0)
    draw.text((260, y), "_____________________________", font=f_body, fill=60)
    draw.text((60,  y + 22), "Date:", font=f_body, fill=0)
    draw.text((260, y + 22), "2026-01-15", font=f_body, fill=0)
    draw.text((500, y), "Supervisor Signature:", font=f_body, fill=0)
    draw.text((700, y), "____________________", font=f_body, fill=60)

    # Footer
    draw.line([(50, 1060), (width - 50, 1060)], fill=80, width=1)
    draw.text((width // 2, 1078),
              "CONTROLLED DOCUMENT — Form EQ-IR-001 Rev.2 | Page 1 of 1",
              font=f_small, fill=80, anchor="mm")

    # ----- Make it look scanned -----
    arr = np.array(img, dtype=np.float32)

    # Add Gaussian noise (simulates scanner grain)
    noise = np.random.normal(0, 4, arr.shape)
    arr = np.clip(arr + noise, 0, 255).astype(np.uint8)
    img = Image.fromarray(arr)

    # Slight blur (scanner focus imperfection)
    img = img.filter(ImageFilter.GaussianBlur(radius=0.6))

    # Slight rotation (paper not perfectly straight on scanner bed)
    angle = random.uniform(0.4, 1.2)
    img = img.rotate(angle, fillcolor=252, expand=False)

    path = RAW / "inspection_scan_001.png"
    img.save(str(path), dpi=(150, 150))
    print(f"  Created: {path.name}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Generating synthetic industrial documents into data/raw/ …\n")

    gen_compressor_c1_manual()
    gen_valve_v12_manual()
    gen_pump_p3_manual()
    gen_sop_process_safety()
    gen_sop_emergency_response()
    gen_maintenance_log()
    gen_incident_reports()
    gen_scanned_inspection()

    print(f"\nDone. {len(list(RAW.iterdir()))} files in {RAW}")
    print("\nContradictions planted:")
    print("  C1 (decay)              Valve-V12      — 3-month interval vs last inspection 2025-11-05")
    print("  C2 (direct)             Compressor-C1  — 150 psi (manual) vs 180 psi (SOP)")
    print("  C3 (direct)             Pump-P3        — 15-min cooldown (manual) vs 5-min (SOP)")
    print("\nSee data/CONTRADICTIONS.md for full detail.")
